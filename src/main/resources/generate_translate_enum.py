import openpyxl
import re
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

I18N_PATHS = [
    "src/main/resources/i18n",
]

OUTPUT_PATH = "src/main/java/com/wis/i18n/Translate.java"
PACKAGE_NAME = "com.wis.i18n"


def to_enum_name(key: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]", "_", key)
    s = re.sub(r"_+", "_", s).upper().strip("_")
    if not s:
        s = "KEY"
    if s[0].isdigit():
        s = "K_" + s
    return s


def main():
    entries = {}
    total_files = 0

    for i18n_dir in I18N_PATHS:
        if not os.path.isdir(i18n_dir):
            print(f"Not found directory: {i18n_dir}")
            continue

        print(f"Scanning directory: {i18n_dir}")
        for filename in os.listdir(i18n_dir):
            if not filename.endswith(".xlsx") or filename.startswith("~$"):
                continue

            total_files += 1
            full_path = os.path.join(i18n_dir, filename)
            print(f"Đọc file: {full_path}")

            try:
                wb = openpyxl.load_workbook(full_path)
                sheet = wb.active

                headers = [str(c.value).strip() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
                key_idx = None
                vi_idx = None

                for i, h in enumerate(headers):
                    if h.lower() in ["key", "id", "code"]:
                        key_idx = i
                    elif h.lower() == "vi_vn":
                        vi_idx = i

                if key_idx is None:
                    print(f"Not found column 'key' in {filename}")
                    continue
                if vi_idx is None:
                    print(f"Not found column 'vi_VN' in {filename}")
                    continue

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    key = row[key_idx]
                    vi_val = row[vi_idx] if vi_idx < len(row) else None

                    if key and str(key).strip():
                        k = str(key).strip()
                        v = str(vi_val).strip() if vi_val else k
                        entries[k] = v

            except Exception as e:
                print(f"Error reading file {filename}: {e}")

    if not entries:
        print("Not found any key in file i18n")
        return

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(f"package {PACKAGE_NAME};\n\n")
        f.write("import lombok.Getter;\nimport lombok.AllArgsConstructor;\nimport lombok.ToString;\n\n")
        f.write("@Getter\n")
        f.write("@AllArgsConstructor\n")
        f.write("@ToString\n")
        f.write("public enum Translate {\n\n")

        items = sorted(entries.items())
        for i, (key, vi_val) in enumerate(items):
            enum_name = to_enum_name(key)
            vi_val = vi_val.replace('"', '\\"')
            sep = ";" if i == len(items) - 1 else ","
            f.write(f'    {enum_name}("{vi_val}"){sep}\n')

        f.write("\n    private final String description;\n")
        f.write("}\n")

    print(f"\n-> Created file: {OUTPUT_PATH}")
    print(f"-> Total keys: {len(entries)} (from {total_files} file Excel)")


if __name__ == "__main__":
    main()
